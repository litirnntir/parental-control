import json
import subprocess
import os
import signal

import openpyxl
import osascript

from QMessages import pop_up_message


# Функция для преобразования секунд в формат часы:минуты:секунды
def format_time(seconds):
    hours = seconds // 3600  # Целочисленное деление на 3600
    minutes = (seconds % 3600) // 60  # Остаток от деления на 3600, затем целочисленное деление на 60
    seconds = (seconds % 3600) % 60  # Остаток от деления на 3600, затем остаток от деления на 60
    return f"{hours}:{minutes:02d}:{seconds:02d}"  # Форматируем строку с нулями слева для минут и секунд


def save_stats_to_file():
    with open("settings.json", "r") as f:
        data = json.load(f)
        directory = data["directory"] + "/stats.txt"

    with open("stats_apps.json", "r") as f:
        stats_apps = json.load(f)

    # Создаем новую книгу excel
    wb = openpyxl.Workbook()

    # Получаем активный лист
    ws = wb.active

    data = stats_apps

    # Проходим по словарю и записываем данные построчно в лист
    row = 1  # Номер строки
    for app, time in data.items():
        ws.cell(row=row, column=1).value = app  # Записываем название приложения в первый столбец
        ws.cell(row=row, column=2).value = format_time(
            time)  # Записываем время во второй столбец в формате часы:минуты:секунды
        row += 1  # Переходим на следующую строку

    # Сохраняем книгу excel в файл
    wb.save("data.xlsx")


def get_from_json(file_name):
    with open(str(file_name), "r") as file:
        # Загружаем данные из файла в переменную data
        data = json.load(file)
    # Присваиваем переменную data словарю self.blocked_apps
    return data


def send_notification(text):
    osascript.run("defaults write com.apple.notificationcenterui bannerTime 2")
    command = f'display notification "{text}" with title "Croak"'
    osascript.run(command)


def apps_list():
    apps = []
    app_path = "/Applications"
    for file in os.listdir(app_path):
        if file.endswith(".app"):
            apps.append(file[:-4])
    return apps


def close_app(app_name):
    try:
        processes = os.popen("ps ax").readlines()
        for process in processes:
            if app_name in process:
                fields = process.split()
                pid = fields[0]
                os.kill(int(pid), signal.SIGTERM)
                print(f"Закрыли приложение {app_name}")
    except:
        pass


def get_open_apps():
    script = 'tell application "System Events" to get name of every process whose background only is false'
    output = subprocess.check_output(['osascript', '-e', script])
    output = output.decode('utf-8').strip().split(', ')
    return output


def get_active_app_name():
    script = """
    tell application "System Events"
        set frontApp to name of first application process whose frontmost is true
    end tell
    return frontApp
    """
    output = subprocess.check_output(["osascript", "-e", script])
    return output.strip().decode("utf-8")
